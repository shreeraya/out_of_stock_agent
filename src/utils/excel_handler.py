import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelHandler:
    """Handles loading input data templates and writing styled, production-grade output Excel reports."""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found at: {file_path}")
            
    def load_inputs(self) -> dict:
        """Loads and validates all necessary input sheets from the Excel file."""
        sheets = {
            "SKU_Metadata": ["SKU", "Description", "Category", "Supplier_ID", "Unit_Cost_USD", "Selling_Price_USD", "Lead_Time_Days"],
            "Inventory_Status": ["SKU", "DC", "Current_Stock_Units", "Safety_Stock_Units", "Reorder_Point_Units", "Reorder_Quantity_Units"],
            "Demand_Forecast": ["SKU", "DC", "Date", "Forecasted_Demand_Units"],
            "Supply_Pipeline": ["SKU", "DC", "Order_ID", "Quantity_Units", "Expected_Delivery_Date", "Status"]
        }
        
        data = {}
        with pd.ExcelFile(self.file_path) as xls:
            for sheet_name, cols in sheets.items():
                if sheet_name not in xls.sheet_names:
                    # Provide an empty DataFrame if Supply_Pipeline is missing, but error on others
                    if sheet_name == "Supply_Pipeline":
                        print("[INFO] Supply_Pipeline sheet missing; assuming no orders in transit.")
                        data[sheet_name] = pd.DataFrame(columns=cols)
                        continue
                    else:
                        raise ValueError(f"Required sheet '{sheet_name}' is missing from {self.file_path}")
                
                df = pd.read_excel(xls, sheet_name=sheet_name)
                
                # Check for missing columns
                missing_cols = [c for c in cols if c not in df.columns]
                if missing_cols:
                    raise ValueError(f"Sheet '{sheet_name}' is missing required columns: {missing_cols}")
                
                # Clean up and select columns
                df = df[cols].copy()
                df["SKU"] = df["SKU"].astype(str).str.strip()
                if "DC" in df.columns:
                    df["DC"] = df["DC"].astype(str).str.strip()
                if "Status" in df.columns:
                    df["Status"] = df["Status"].astype(str).str.strip()
                    
                data[sheet_name] = df
                
        return data

    def write_analysis_report(self, output_path: str, input_data: dict, oos_risks: list, rca_results: list, mitigation_results: list):
        """Creates a beautiful, executive-ready Excel workbook compiling inputs and analysis."""
        wb = openpyxl.Workbook()
        
        # Color Palette - Executive Navy/Slate theme
        HEADER_FILL = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid") # Dark Navy
        ZEBRA_FILL = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid") # Light Blue-Gray Zebra
        WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Severity / Priority Fills & Fonts
        HIGH_SEVERITY_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid") # Pastel Red
        HIGH_SEVERITY_FONT = Font(name="Segoe UI", size=10, bold=True, color="78281F")
        
        MEDIUM_SEVERITY_FILL = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid") # Pastel Orange
        MEDIUM_SEVERITY_FONT = Font(name="Segoe UI", size=10, bold=True, color="7E5109")
        
        LOW_SEVERITY_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid") # Pastel Green
        LOW_SEVERITY_FONT = Font(name="Segoe UI", size=10, bold=True, color="196F3D")
        
        # General Fonts
        FONT_FAMILY = "Segoe UI"
        header_font = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
        title_font = Font(name=FONT_FAMILY, size=16, bold=True, color="1A365D")
        subtitle_font = Font(name=FONT_FAMILY, size=10, italic=True, color="7F8C8D")
        bold_font = Font(name=FONT_FAMILY, size=10, bold=True, color="1A365D")
        regular_font = Font(name=FONT_FAMILY, size=10, color="2C3E50")
        
        # Card Fonts (for dashboard)
        card_label_font = Font(name=FONT_FAMILY, size=9, bold=True, color="7F8C8D")
        card_val_font = Font(name=FONT_FAMILY, size=18, bold=True, color="1A365D")
        
        # Alignments
        left_align = Alignment(horizontal="left", vertical="center")
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        wrap_left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Borders
        thin_side = Side(border_style="thin", color="D5DBDB")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        double_bottom_side = Side(border_style="double", color="1A365D")
        thick_bottom_side = Side(border_style="medium", color="1A365D")
        header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom_side)
        card_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Helper function to style normal sheets
        def add_styled_sheet(df, name, is_dashboard=False):
            ws = wb.create_sheet(title=name)
            ws.views.sheetView[0].showGridLines = True
            
            start_row = 2
            
            if is_dashboard:
                # Add Dashboard Title
                ws.cell(row=2, column=1, value="StockSentinel - Executive OOS Risk Dashboard").font = title_font
                ws.cell(row=3, column=1, value=f"Generated on {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}. AI-Assisted Inventory Integrity Engine.").font = subtitle_font
                
                # Create KPI Summary Cards
                total_skus = len(input_data["SKU_Metadata"]["SKU"].unique())
                total_oos_events = len(df)
                high_severity = len(df[df["Severity_Level"] == "High"]) if "Severity_Level" in df.columns else 0
                
                # Card 1: SKUs Scanned
                ws.merge_cells("B5:C5")
                ws.merge_cells("B6:C6")
                ws.cell(row=5, column=2, value="TOTAL SKUS SCANNED").font = card_label_font
                ws.cell(row=5, column=2).alignment = center_align
                ws.cell(row=6, column=2, value=total_skus).font = card_val_font
                ws.cell(row=6, column=2).alignment = center_align
                ws.cell(row=6, column=2).number_format = "#,##0"
                
                # Card 2: Stockout Risks Detected
                ws.merge_cells("E5:F5")
                ws.merge_cells("E6:F6")
                ws.cell(row=5, column=5, value="OOS RISKS DETECTED").font = card_label_font
                ws.cell(row=5, column=5).alignment = center_align
                ws.cell(row=6, column=5, value=total_oos_events).font = card_val_font
                ws.cell(row=6, column=5).alignment = center_align
                ws.cell(row=6, column=5).number_format = "#,##0"
                
                # Card 3: High Severity
                ws.merge_cells("H5:I5")
                ws.merge_cells("H6:I6")
                ws.cell(row=5, column=8, value="HIGH SEVERITY RISKS").font = card_label_font
                ws.cell(row=5, column=8).alignment = center_align
                ws.cell(row=6, column=8, value=high_severity).font = card_val_font
                ws.cell(row=6, column=8).alignment = center_align
                ws.cell(row=6, column=8).number_format = "#,##0"
                
                # Border & fill KPI Cards
                card_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                for r in [5, 6]:
                    for c in [2, 3, 5, 6, 8, 9]:
                        cell = ws.cell(row=r, column=c)
                        cell.fill = card_fill
                        cell.border = card_border
                        
                start_row = 9 # Table starts further down for dashboard
                
            # Write Headers
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = HEADER_FILL
                cell.alignment = center_align
                cell.border = header_border
            
            ws.row_dimensions[start_row].height = 28
            
            # Write Data
            for row_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
                is_zebra = (row_idx % 2 == 0)
                row_fill = ZEBRA_FILL if is_zebra else WHITE_FILL
                
                # Check for severity/priority in rows to apply special conditional styling
                custom_row_styling = None
                severity_val = None
                
                if "Severity_Level" in headers:
                    severity_val = row[headers.index("Severity_Level")]
                elif "Priority_Level" in headers:
                    severity_val = row[headers.index("Priority_Level")]
                    
                for col_idx, value in enumerate(row, start=1):
                    # Robust type cleaning: convert list or dictionary structures to string
                    if isinstance(value, list):
                        value = "\n".join(str(v) for v in value)
                    elif isinstance(value, dict):
                        import json
                        value = json.dumps(value)
                        
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = regular_font
                    cell.fill = row_fill
                    cell.border = thin_border
                    
                    header_name = headers[col_idx - 1]
                    
                    # Alignment and formats
                    if "SKU" in header_name or "DC" in header_name or "ID" in header_name or "Status" in header_name:
                        cell.alignment = center_align
                    elif "Date" in header_name:
                        cell.alignment = center_align
                    elif "Cost" in header_name or "Price" in header_name or "USD" in header_name:
                        cell.number_format = "$#,##0.00"
                        cell.alignment = right_align
                    elif "Units" in header_name or "Stock" in header_name or "Impact" in header_name or "Days" in header_name:
                        cell.number_format = "#,##0"
                        cell.alignment = right_align
                    elif "Action" in header_name or "Steps" in header_name or "Reasoning" in header_name or "Cause" in header_name:
                        cell.alignment = wrap_left_align
                    else:
                        cell.alignment = left_align
                        
                    # Apply specific severity cell color fills
                    if header_name in ["Severity_Level", "Priority_Level"] and value in ["High", "Medium", "Low"]:
                        if value == "High":
                            cell.fill = HIGH_SEVERITY_FILL
                            cell.font = HIGH_SEVERITY_FONT
                        elif value == "Medium":
                            cell.fill = MEDIUM_SEVERITY_FILL
                            cell.font = MEDIUM_SEVERITY_FONT
                        elif value == "Low":
                            cell.fill = LOW_SEVERITY_FILL
                            cell.font = LOW_SEVERITY_FONT
                            
                ws.row_dimensions[row_idx].height = 22
                
            # Column auto-fitting
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row >= start_row and cell.value is not None:
                        val_str = str(cell.value)
                        # Truncate extremely long comments so column width doesn't blow out
                        if len(val_str) > 50:
                            val_str = val_str[:47] + "..."
                        max_len = max(max_len, len(val_str))
                
                header_val = ws.cell(row=start_row, column=col[0].column).value
                header_len = len(str(header_val)) if header_val else 0
                ws.column_dimensions[col_letter].width = max(max_len, header_len) + 4
                
        # 1. Output Dashboard
        df_oos = pd.DataFrame(oos_risks)
        if df_oos.empty:
            df_oos = pd.DataFrame(columns=["SKU", "DC", "Date_of_OOS", "Days_Until_OOS", "Current_Stock", "Projected_Stock_on_OOS_Date", "Severity_Level"])
        add_styled_sheet(df_oos, "OOS_Risk_Dashboard", is_dashboard=True)
        
        # 2. Output RCA
        df_rca = pd.DataFrame(rca_results)
        if df_rca.empty:
            df_rca = pd.DataFrame(columns=["SKU", "DC", "Date_of_OOS", "Days_Until_OOS", "Primary_Root_Cause", "Secondary_Factors", "Narrative_Reasoning"])
        add_styled_sheet(df_rca, "Root_Cause_Analysis")
        
        # 3. Output Mitigation Recommendations
        df_mitigation = pd.DataFrame(mitigation_results)
        if df_mitigation.empty:
            df_mitigation = pd.DataFrame(columns=["SKU", "DC", "Date_of_OOS", "Recommended_Action", "Action_Steps", "Inventory_Impact_Units", "Estimated_Cost_USD", "Priority_Level"])
        add_styled_sheet(df_mitigation, "Mitigation_Recommendations")
        
        # 4. Copy Input sheets for reference
        for sheet_name, df_in in input_data.items():
            add_styled_sheet(df_in, f"In_{sheet_name}")
            
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
            
        try:
            wb.save(output_path)
            print(f"[SUCCESS] Styled Excel analysis report saved as: {output_path}")
        except PermissionError:
            import datetime
            from pathlib import Path
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            path_obj = Path(output_path)
            alternative_path = str(path_obj.parent / f"{path_obj.stem}_{timestamp}{path_obj.suffix}")
            print(f"[WARNING] Permission denied to write to '{output_path}'. The file is likely open in Microsoft Excel.")
            print(f"[STATUS] Saving the report under an alternative timestamped name: '{alternative_path}'")
            wb.save(alternative_path)
            print(f"[SUCCESS] Styled Excel analysis report saved as: {alternative_path}")
