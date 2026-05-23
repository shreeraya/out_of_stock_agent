import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_styled_excel():
    # Define color palette (Executive Slate Theme)
    HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid") # Dark Navy/Slate
    ZEBRA_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid") # Very light gray
    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Text fonts
    FONT_FAMILY = "Segoe UI"
    header_font = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
    title_font = Font(name=FONT_FAMILY, size=16, bold=True, color="2C3E50")
    subtitle_font = Font(name=FONT_FAMILY, size=10, italic=True, color="7F8C8D")
    bold_font = Font(name=FONT_FAMILY, size=10, bold=True, color="2C3E50")
    regular_font = Font(name=FONT_FAMILY, size=10, color="2C3E50")
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    # Borders
    thin_border_side = Side(border_style="thin", color="D5DBDB")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom_side = Side(border_style="medium", color="2C3E50")
    header_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
    
    # Prepare Mock Data (Weekly Granularity)
    
    # 1. SKU Metadata (Lead times in weeks)
    sku_data = [
        {"SKU": "SKU-001", "Description": "Premium Wireless Headphones", "Category": "Electronics", "Supplier_ID": "SUPP-101", "Unit_Cost_USD": 45.00, "Selling_Price_USD": 99.00, "Lead_Time_Weeks": 1.0, "Lead_Time_StdDev_Weeks": 0.2},
        {"SKU": "SKU-002", "Description": "Ergonomic Office Chair", "Category": "Furniture", "Supplier_ID": "SUPP-202", "Unit_Cost_USD": 75.00, "Selling_Price_USD": 180.00, "Lead_Time_Weeks": 2.0, "Lead_Time_StdDev_Weeks": 0.4},
        {"SKU": "SKU-003", "Description": "Organic Matcha Tea Powder 100g", "Category": "Pantry", "Supplier_ID": "SUPP-303", "Unit_Cost_USD": 8.50, "Selling_Price_USD": 22.00, "Lead_Time_Weeks": 2.0, "Lead_Time_StdDev_Weeks": 0.5},
        {"SKU": "SKU-004", "Description": "Smart Fitness Tracker Band", "Category": "Electronics", "Supplier_ID": "SUPP-101", "Unit_Cost_USD": 20.00, "Selling_Price_USD": 49.00, "Lead_Time_Weeks": 1.0, "Lead_Time_StdDev_Weeks": 0.2},
        {"SKU": "SKU-005", "Description": "Biodegradable Bamboo Coffee Mug", "Category": "Housewares", "Supplier_ID": "SUPP-404", "Unit_Cost_USD": 3.00, "Selling_Price_USD": 12.00, "Lead_Time_Weeks": 1.0, "Lead_Time_StdDev_Weeks": 0.3}
    ]
    df_sku = pd.DataFrame(sku_data)
    
    # 2. Inventory Status (safety/reorder quantities in units, demand stddev as weekly)
    inventory_data = [
        {"SKU": "SKU-001", "DC": "DC-EAST", "Current_Stock_Units": 200, "Safety_Stock_Units": 50, "Reorder_Point_Units": 80, "Reorder_Quantity_Units": 150, "Weekly_Demand_StdDev_Units": 12.5},
        {"SKU": "SKU-002", "DC": "DC-EAST", "Current_Stock_Units": 120, "Safety_Stock_Units": 40, "Reorder_Point_Units": 60, "Reorder_Quantity_Units": 100, "Weekly_Demand_StdDev_Units": 18.0},
        {"SKU": "SKU-003", "DC": "DC-WEST", "Current_Stock_Units": 50, "Safety_Stock_Units": 30, "Reorder_Point_Units": 60, "Reorder_Quantity_Units": 120, "Weekly_Demand_StdDev_Units": 15.0},
        {"SKU": "SKU-004", "DC": "DC-EAST", "Current_Stock_Units": 10, "Safety_Stock_Units": 40, "Reorder_Point_Units": 50, "Reorder_Quantity_Units": 100, "Weekly_Demand_StdDev_Units": 10.0},
        {"SKU": "SKU-005", "DC": "DC-EAST", "Current_Stock_Units": 15, "Safety_Stock_Units": 25, "Reorder_Point_Units": 35, "Reorder_Quantity_Units": 80, "Weekly_Demand_StdDev_Units": 8.0},
        {"SKU": "SKU-005", "DC": "DC-WEST", "Current_Stock_Units": 300, "Safety_Stock_Units": 40, "Reorder_Point_Units": 80, "Reorder_Quantity_Units": 150, "Weekly_Demand_StdDev_Units": 5.0} # Transfer source
    ]
    df_inventory = pd.DataFrame(inventory_data)
    
    # 3. Demand Forecast (12-week projection starting from next Monday)
    today = datetime.date.today()
    next_monday = today + datetime.timedelta(days=(7 - today.weekday()) % 7)
    if next_monday == today:
        next_monday = today + datetime.timedelta(days=7)
        
    demand_records = []
    
    for week in range(12):
        current_date = next_monday + datetime.timedelta(weeks=week)
        date_str = current_date.strftime("%Y-%m-%d")
        
        # SKU-001: DC-EAST (Steady demand, well stocked)
        demand_records.append({"SKU": "SKU-001", "DC": "DC-EAST", "Week_Start_Date": date_str, "Forecasted_Demand_Units": 60})
        
        # SKU-002: DC-EAST (Promotional spike starting Week 4)
        # Steady at 40/week, then spikes to 200/week in Week 4 and Week 5, then back to 40/week
        demand = 40
        if 4 <= week <= 5:
            demand = 200
        demand_records.append({"SKU": "SKU-002", "DC": "DC-EAST", "Week_Start_Date": date_str, "Forecasted_Demand_Units": demand})
        
        # SKU-003: DC-WEST (Steady demand, but delayed PO causes stockout)
        # Demand is 50/week
        demand_records.append({"SKU": "SKU-003", "DC": "DC-WEST", "Week_Start_Date": date_str, "Forecasted_Demand_Units": 50})
        
        # SKU-004: DC-EAST (Low initial stock, steady demand)
        # Demand is 40/week
        demand_records.append({"SKU": "SKU-004", "DC": "DC-EAST", "Week_Start_Date": date_str, "Forecasted_Demand_Units": 40})
        
        # SKU-005: DC-EAST & DC-WEST (Stock transfer scenario)
        # DC-EAST demand: 25/week. Stock starts at 15 -> runs dry Week 1
        demand_records.append({"SKU": "SKU-005", "DC": "DC-EAST", "Week_Start_Date": date_str, "Forecasted_Demand_Units": 25})
        # DC-WEST demand: 20/week. Stock starts at 300 -> stays completely healthy
        demand_records.append({"SKU": "SKU-005", "DC": "DC-WEST", "Week_Start_Date": date_str, "Forecasted_Demand_Units": 20})
        
    df_demand = pd.DataFrame(demand_records)
    
    # 4. Supply Pipeline (snapped to week-start Mondays)
    pipeline_data = [
        # SKU-001: DC-EAST (Order placed, arriving on Week 4)
        {"SKU": "SKU-001", "DC": "DC-EAST", "Order_ID": "PO-991", "Quantity_Units": 150, "Expected_Delivery_Week_Start": (next_monday + datetime.timedelta(weeks=4)).strftime("%Y-%m-%d"), "Status": "In Transit"},
        
        # SKU-002: DC-EAST (Emergency PO arriving Week 8 - too late for the promotion spike in Week 4!)
        {"SKU": "SKU-002", "DC": "DC-EAST", "Order_ID": "PO-992", "Quantity_Units": 100, "Expected_Delivery_Week_Start": (next_monday + datetime.timedelta(weeks=8)).strftime("%Y-%m-%d"), "Status": "Placed"},
        
        # SKU-003: DC-WEST (Late supplier delivery - originally expected Week 1, but delayed to Week 6!)
        {"SKU": "SKU-003", "DC": "DC-WEST", "Order_ID": "PO-993", "Quantity_Units": 120, "Expected_Delivery_Week_Start": (next_monday + datetime.timedelta(weeks=6)).strftime("%Y-%m-%d"), "Status": "Delayed"},
    ]
    df_pipeline = pd.DataFrame(pipeline_data)
    
    # Create excel workbook
    wb = openpyxl.Workbook()
    
    # 1. Sheet README
    ws_readme = wb.active
    ws_readme.title = "README"
    ws_readme.views.sheetView[0].showGridLines = True
    
    ws_readme["A2"] = "StockSentinel - Weekly Multi-Agent OOS Guard"
    ws_readme["A2"].font = title_font
    ws_readme["A3"] = "Input Template & Operating Instructions (Weekly Granularity)"
    ws_readme["A3"].font = subtitle_font
    
    instructions = [
        "",
        "Welcome to the weekly-focused StockSentinel Multi-Agent Out-of-Stock (OOS) Predictor framework.",
        "This spreadsheet serves as the template for loading your supply chain data at a weekly granularity.",
        "The Python engine simulates weekly inventory logs, calculates volatility risks, and invokes collaborative",
        "AI agents to diagnose root causes and recommend actionable operational mitigation strategies.",
        "",
        "DIRECTIONS FOR USE:",
        "1. Populate the 4 orange-tabbed sheets with your actual supply chain data.",
        "   - SKU_Metadata: Basic information, unit costs, and supplier lead times in WEEKS.",
        "   - Inventory_Status: Current inventory snapshot, safety stock, and weekly demand variability.",
        "   - Demand_Forecast: Forecasted unit sales aggregated by Week_Start_Date (Mondays) for a 12-week horizon.",
        "   - Supply_Pipeline: Outstanding POs, expected delivery week start dates, and status.",
        "2. Configure your OpenAI API Key in a '.env' file in the project folder.",
        "3. Run the analysis engine by executing: python run_analysis.py",
        "4. A beautifully styled report ('oos_analysis_report.xlsx') will be generated, containing the",
        "   executive OOS Risk Dashboard (showing Week of OOS, Weeks Until OOS, and Probability Scores),",
        "   Root Cause Diagnoses, and Action Directives.",
        "",
        "MOCK SCENARIOS DEMONSTRATED IN THIS WEEKLY DATASET:",
        "- SKU-001 (DC-EAST): Steady demand & supply pipeline. Expected stock stays completely healthy.",
        "- SKU-002 (DC-EAST): Promotional Demand Spike on Weeks 4-5. Stockout is predicted in Week 4 because lead time (2 weeks)",
        "  prevents replenishment from arriving in time. Needs expedited shipping or earlier ordering.",
        "- SKU-003 (DC-WEST): Late Supplier Delivery. PO-993 is delayed from Week 1 to Week 6.",
        "  Inventory runs dry in Week 2 and stays out-of-stock for 4 weeks. Needs urgent supplier escalation.",
        "- SKU-004 (DC-EAST): Low Initial Stock. Starting inventory is 10 units with no PO in transit. Runs out in Week 1.",
        "- SKU-005 (DC-EAST): Distribution imbalance. DC-EAST runs dry in Week 1, but DC-WEST has excess safety stock.",
        "  Perfect candidate for a Plant-to-Plant Stock Transport Order (STO) transfer."
    ]
    
    for row_idx, line in enumerate(instructions, start=5):
        cell = ws_readme.cell(row=row_idx, column=1, value=line)
        cell.font = regular_font
        if "DIRECTIONS FOR USE:" in line or "MOCK SCENARIOS DEMONSTRATED:" in line:
            cell.font = bold_font
            
    ws_readme.column_dimensions["A"].width = 110
    
    # Styling helper for pandas dataframes
    def write_sheet(df, title, tab_color=None):
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True
        if tab_color:
            ws.sheet_properties.tabColor = tab_color
            
        # Write headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = HEADER_FILL
            cell.alignment = center_align
            cell.border = header_border
        
        # Write data rows
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            is_zebra = (row_idx % 2 == 0)
            row_fill = ZEBRA_FILL if is_zebra else WHITE_FILL
            
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = regular_font
                cell.fill = row_fill
                cell.border = thin_border
                
                # Format cells based on headers/data types
                header_name = headers[col_idx - 1]
                if "SKU" in header_name or "DC" in header_name or "ID" in header_name or "Status" in header_name:
                    cell.alignment = center_align
                elif "Date" in header_name or "Week_Start" in header_name:
                    cell.alignment = center_align
                elif "Cost" in header_name or "Price" in header_name:
                    cell.number_format = "$#,##0.00"
                    cell.alignment = right_align
                elif "Units" in header_name or "Stock" in header_name or "Point" in header_name or "Quantity" in header_name or "Weeks" in header_name or "Demand" in header_name:
                    cell.number_format = "#,##0.0" if "StdDev" in header_name or "Weeks" in header_name else "#,##0"
                    cell.alignment = right_align
                else:
                    cell.alignment = left_align
                    
        # Adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    # Special format length checking
                    if isinstance(cell.value, float) and ("Cost" in headers[col[0].column-1] or "Price" in headers[col[0].column-1]):
                        val_str = f"${cell.value:,.2f}"
                    elif isinstance(cell.value, int) and ("Units" in headers[col[0].column-1] or "Stock" in headers[col[0].column-1]):
                        val_str = f"{cell.value:,}"
                    else:
                        val_str = str(cell.value)
                    max_len = max(max_len, len(val_str))
            
            header_val = ws.cell(row=1, column=col[0].column).value
            header_len = len(str(header_val)) if header_val else 0
            ws.column_dimensions[col_letter].width = max(max_len, header_len) + 4
            
        ws.row_dimensions[1].height = 26
        
    # Write sheets with specific tab colors (orange for inputs)
    INPUT_TAB_COLOR = "E67E22" # Soft Orange
    write_sheet(df_sku, "SKU_Metadata", INPUT_TAB_COLOR)
    write_sheet(df_inventory, "Inventory_Status", INPUT_TAB_COLOR)
    write_sheet(df_demand, "Demand_Forecast", INPUT_TAB_COLOR)
    write_sheet(df_pipeline, "Supply_Pipeline", INPUT_TAB_COLOR)
    
    # Save Workbook
    filename = "input_template.xlsx"
    wb.save(filename)
    print(f"[SUCCESS] Beautiful Excel input template saved as: {filename}")

if __name__ == "__main__":
    create_styled_excel()
